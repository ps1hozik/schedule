import os
import logging
import re
from dataclasses import dataclass, field
from typing import List, Optional, Tuple
from datetime import datetime

from tqdm import tqdm

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from settings import DATA_FOLDER


@dataclass(frozen=True)
class Group:
    faculty: str
    name: str
    course: int
    specialty: str
    form: str
    column_range: range = field(compare=False)
    subgroups: Tuple[str, ...] = field(default_factory=tuple)


def parse_date_custom(date_str):
    month_mapping = {
        "января": "01",
        "февраля": "02",
        "марта": "03",
        "апреля": "04",
        "мая": "05",
        "июня": "06",
        "июля": "07",
        "августа": "08",
        "сентября": "09",
        "октября": "10",
        "ноября": "11",
        "декабря": "12",
    }
    try:
        date_list = date_str.strip().split()[:3]
        date_list[1] = month_mapping[date_list[1]]
        date_str = ".".join(date_list)
    except Exception as e:
        pass

    parsed_date = datetime.strptime(date_str, "%d.%m.%Y").date()
    return parsed_date


@dataclass(frozen=True)
class Pair:
    week_day: str
    date: datetime.date
    number: int
    subgroup: str
    specialty: str
    teacher: Optional[str] = None
    auditorium: Optional[str] = None
    name: Optional[str] = None

    def __post_init__(self):
        object.__setattr__(self, "date", parse_date_custom(self.date))

    @staticmethod
    def get_week_day(ws, row, week_day) -> str:
        week_day_column = ws.cell(row=row, column=Schedule.start_column - 3).value
        if week_day == "" or (
            week_day_column is not None and week_day != week_day_column
        ):
            return week_day_column
        return week_day

    @staticmethod
    def get_pair_date(ws, row, pair_date) -> str:
        pair_date_column = ws.cell(row=row, column=Schedule.start_column - 2).value
        if pair_date == "" or (
            pair_date_column is not None and pair_date != pair_date_column
        ):
            return pair_date_column
        return pair_date

    @staticmethod
    def get_pair_number(ws, row) -> int:
        number_column: str = ws.cell(row=row, column=Schedule.start_column - 1).value
        return (
            number_column
            if number_column.isdigit()
            else ws.cell(row=row - 1, column=Schedule.start_column - 1).value
        )


@dataclass(frozen=True)
class ExamCredit:
    week_day: str
    date: datetime.date
    subgroup: str
    specialty: str
    name: Optional[str] = None
    teacher: Optional[str] = None
    auditorium: Optional[str] = None
    time: Optional[str] = None

    def __post_init__(self):
        object.__setattr__(self, "date", parse_date_custom(self.date))


class Schedule:
    start_row = 0
    start_column = 0


def _find_merged_range(cell, merged_ranges) -> MergedCellRange | None:
    for merged_range in merged_ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None


def _get_groups(
    worksheet: Worksheet,
    merged_ranges: set[CellRange],
    faculty: str,
    form: str,
) -> Tuple[List[Group], str]:
    groups = []
    pair_type = "обыч"
    for row in worksheet.iter_rows():
        for cell in row:
            cell_value = str(cell.value).lower()
            if (
                "зачет" in cell_value.lower()
                or "зачёт" in cell_value.lower()
                or "test" in cell_value.lower()
            ):
                pair_type = "зач"
            elif "экзамен" in cell_value.lower() or "exam" in cell_value.lower():
                pair_type = "экз"
            if re.match(r"\d (курс|year)", cell_value):
                subgroups = []
                course = int(cell_value[0])
                specialty = worksheet.cell(row=cell.row + 1, column=cell.column).value
                group_name = worksheet.cell(row=cell.row + 2, column=cell.column).value
                if not all([course, specialty, group_name]):
                    continue
                subgroups_row = cell.row + 3
                merged_range = _find_merged_range(cell, merged_ranges)
                max_column = merged_range.max_col if merged_range else cell.column
                for col_idx in range(cell.column, max_column + 1):
                    subgroup = str(
                        worksheet.cell(row=subgroups_row, column=col_idx).value
                    )
                    if group_name in subgroup:
                        subgroups.append(subgroup)
                if not subgroups:
                    continue
                group = Group(
                    column_range=range(cell.column, max_column + 1),
                    faculty=str(faculty),
                    name=str(group_name),
                    course=course,
                    specialty=str(specialty),
                    subgroups=tuple(subgroups),
                    form=form,
                )
                groups.append(group)
                if not Schedule.start_row:
                    Schedule.start_row = subgroups_row + 2
                    Schedule.start_column = (
                        merged_range.min_col if merged_range else cell.column
                    )
    return groups, pair_type


def _get_pairs(
    ws: Worksheet, groups: List[Group], merged_ranges: set[CellRange]
) -> List[Pair]:
    pairs = []
    for group in groups:
        for idx, column in enumerate(group.column_range):
            row = Schedule.start_row
            week_day = ""
            pair_date = ""
            while ws.cell(row=row, column=Schedule.start_column - 1).value:

                number = Pair.get_pair_number(ws=ws, row=row)
                week_day = Pair.get_week_day(ws=ws, row=row, week_day=week_day)
                pair_date = Pair.get_pair_date(ws=ws, row=row, pair_date=pair_date)

                if isinstance(ws.cell(row=row, column=column), MergedCell):
                    merged_range = _find_merged_range(
                        ws.cell(row=row, column=column), merged_ranges
                    )
                    name = ws.cell(row=row, column=merged_range.min_col).value
                    teacher = ws.cell(row=row + 1, column=merged_range.min_col).value
                    auditorium = ws.cell(row=row + 2, column=merged_range.min_col).value
                else:
                    name = ws.cell(row=row, column=column).value
                    teacher = ws.cell(row=row + 1, column=column).value
                    auditorium = ws.cell(row=row + 2, column=column).value
                subgroup = group.subgroups[idx]
                pairs.append(
                    Pair(
                        week_day=week_day,
                        date=pair_date,
                        number=number,
                        name=name,
                        teacher=teacher,
                        auditorium=auditorium,
                        subgroup=subgroup,
                        specialty=group.specialty,
                    )
                )
                row += 3
                if not ws.cell(row=row, column=Schedule.start_column - 1).value:
                    row += 1
    return pairs


def _get_exam_credit(
    ws: Worksheet, groups: List[Group], merged_ranges: set[CellRange]
) -> List[ExamCredit]:
    pairs = []
    for group in groups:
        for idx, column in enumerate(group.column_range):
            row = Schedule.start_row
            while ws.cell(row=row, column=Schedule.start_column - 1).value:

                date_cell = ws.cell(
                    row=row, column=Schedule.start_column - 1
                ).value.split(maxsplit=1)

                week_day = date_cell[1]
                pair_date = date_cell[0]

                if isinstance(ws.cell(row=row, column=column), MergedCell):
                    merged_range = _find_merged_range(
                        ws.cell(row=row, column=column), merged_ranges
                    )
                    data = ws.cell(row=row, column=merged_range.min_col).value
                else:
                    data = ws.cell(row=row, column=column).value
                subgroup = group.subgroups[idx]
                try:
                    name, teacher, auditorium_time = data.rsplit("\n", maxsplit=2)
                    auditorium, time = auditorium_time.rsplit(",", maxsplit=1)
                    time = time.strip()
                except Exception as e:  # ValueError or AttributeError
                    name = teacher = auditorium = time = None
                pairs.append(
                    ExamCredit(
                        week_day=week_day,
                        date=pair_date,
                        name=name,
                        teacher=teacher,
                        auditorium=auditorium,
                        time=time,
                        subgroup=subgroup,
                        specialty=group.specialty,
                    )
                )
                row += 1
    return pairs


def _get_all_files() -> List[Tuple[str, str, str]]:
    files_with_details = []
    for root, _, files in os.walk(DATA_FOLDER):
        if files:
            faculty, form = root.split(os.sep)[1:]
            for file in files:
                files_with_details.append((faculty, form, file))
    return files_with_details


def get_parsed_data() -> Tuple[List[Group], List[Pair], List[ExamCredit]]:
    all_pairs = []
    all_groups = []
    all_exams_credits = []

    files_with_details = _get_all_files()

    format_description = lambda data: data[:47] + "..." if len(data) > 47 else data

    with tqdm(files_with_details, desc="Обработка данных", ncols=150) as progress_bar:
        for faculty, form, file in progress_bar:
            Schedule.start_row = 0
            Schedule.start_column = 0

            description = format_description(str(file))
            progress_bar.set_description(f"Обработка '{description:<50}'")

            file_path = os.path.join("data", faculty, form, file)

            workbook = load_workbook(filename=file_path)
            worksheet = workbook.active
            merged_ranges = worksheet.merged_cells.ranges

            groups, pair_type = _get_groups(worksheet, merged_ranges, faculty, form)
            if not groups:
                continue
            all_groups.extend(groups)
            try:
                if pair_type == "обыч":
                    pairs = _get_pairs(worksheet, groups, merged_ranges)
                    all_pairs.extend(pairs)
                else:
                    pairs = _get_exam_credit(worksheet, groups, merged_ranges)
                    all_exams_credits.extend(pairs)
            except Exception as e:
                print("")
                logging.error(f"{e}\nfile : {file_path}\n")

    unique = lambda data: list(set(data))

    return (
        unique(all_groups),
        unique(all_pairs),
        unique(all_exams_credits),
    )


if __name__ == "__main__":
    groups, pairs, exams_credits = get_parsed_data()
    print(f"Найдено групп: {len(groups)}")
    print(f"Найдено пар: {len(pairs)}")
    print(f"Найдено экзаменов и зачётов: {len(exams_credits)}")
